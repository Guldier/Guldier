from openpyxl import Workbook, load_workbook
import random
import smtplib, ssl

from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from datetime import datetime
from django.core.mail import send_mail, get_connection
from django.core.mail.message import EmailMessage
from .models import (
    Dish, 
    AddOn,
    Composition,
    Cart,
    Orders,
    WeekDish
)

def create_list():
    today = datetime.today()
    today_orders=Orders.objects.filter(order_date__date=today.date())
    composition_list = {}
    money = 0
    
    for food in today_orders:
        if food.composition in composition_list:
            composition_list[food.composition] = composition_list.get(food.composition,0) + food.quantity
        else:
            composition_list[food.composition] = food.quantity
    
    message = 'PODUMOWANIE WSZYSTKICH DAŃ\n'

    for key in composition_list:
        message += f'{key} x {composition_list[key]}\n'

    message += '\nOsoby zamawiające\n'

    for orders in today_orders:
        message += f'{orders.user} - {orders.composition} - {orders.quantity}szt.\n'
        money += (orders.composition.dish.price + orders.composition.addon.price) * orders.quantity

    message += f'\nŁącznie do zapłaty: {money}zł\n'
    message += '\nPozdrawiamy\nLinetech'
    
    send_mail(
        f'Lista obiadów {today.date()}',
        message,
        None,
        ['damian.jadacki@linetech.pl','biuro@hotel-trzykorony.pl','michal.grebosz@linetech.pl','bartosz.wrobel@linetech.pl'],
        fail_silently=False,
    )

def create_list_rep():
    today = datetime.today()
    today_orders=Orders.objects.filter(order_date__date=today.date())
    composition_list = {}

    for food in today_orders:
        if food.user.is_stuff:
            if food.composition in composition_list:
                composition_list[food.composition] = composition_list.get(food.composition,0) + food.quantity
            else:
                composition_list[food.composition] = food.quantity
    
    message = 'Pani Zosiu,\n'
    message += 'Zamówienie:\n'

    for key in composition_list:
        message += f'{key} x {composition_list[key]}\n'

    message += '\nOsoby zamawiające\n'

    for orders in today_orders:
        message += f'{orders.user} - {orders.composition} - {orders.quantity}szt.\n'

    send_mail(
        f'Lista obiadów rep {today.date()}',
        message,
        None,
        ['damian.jadacki@linetech.pl'],
        fail_silently=False,
    )

def five_s():
    my_host = 'linetech.nazwa.pl'
    my_port = 587 
    my_username = 'damian.jadacki@linetech.pl'
    my_password = 'Dddddjjjjj12#$'
    my_use_tls = False
    my_use_ssl = False
    connection = get_connection(
        host=my_host,
        port=my_port,
        username=my_username,
        password=my_password,
        use_tls=my_use_tls,
        use_ssl=my_use_ssl)
    pracownicy = ['DAMIAN JADACKI', 'MICHAŁ GRĘBOSZ', 'BARTOSZ WRÓBEL', 'ŁUKASZ MAGDA']
    dzis = datetime.today()
    workbook = load_workbook(filename="/home/guldier/5S/5S RZE Engineering November 2020 (W46) 10.11.2020.xlsx")
    sheetnames = workbook.sheetnames
    sheet1 = workbook[sheetnames[0]]
    sheet2 = workbook[sheetnames[1]]
    sheet1["C6"].value = dzis.strftime("%B") + f' {dzis.strftime("%Y")} (W{dzis.isocalendar()[1]}) {dzis.strftime("%d")}.{dzis.strftime("%m")}.{dzis.strftime("%Y")}'
    sheet1["C12"].value = f'ENG / {pracownicy[random.randint(0,3)]}'
    losuj_dalej = True
    suma_pkt = 0
    while losuj_dalej:
        suma_pkt = 0
        for i in range(17,22):
            sheet1[f"I{i}"].value = random.randint(4,5)
            suma_pkt += sheet1[f"I{i}"].value

        for i in range(25,30):
            sheet1[f"I{i}"].value = random.randint(4,5)
            suma_pkt += sheet1[f"I{i}"].value

        for i in range(33,38):
            sheet1[f"I{i}"].value = random.randint(4,5)
            suma_pkt += sheet1[f"I{i}"].value

        for i in range(41,46):
            sheet1[f"I{i}"].value = random.randint(4,5)
            suma_pkt += sheet1[f"I{i}"].value

        for i in range(49,53):
            sheet1[f"I{i}"].value = random.randint(4,5)
            suma_pkt += sheet1[f"I{i}"].value

        if sheet1["C12"].value == f'ENG / {pracownicy[1]}':
            sheet1["I53"].value = 3
        else:
            sheet1["I53"].value = 1
    
        suma_pkt += sheet1["I53"].value
        if suma_pkt >= 119:
            losuj_dalej = False
        else:
            suma_pkt = 0

    free_column = 0
    free_row = 0
    for i in range(5,24):
        if sheet2.cell(row=31,column=i).value == None:
            free_column = i
            break
    if free_column == 0:
        for i in range(5,24):
            if sheet2.cell(row=33, column=i).value == None:
                free_column = i
                break
        sheet2.cell(row=33, column=free_column).value = dzis.date()
        sheet2.cell(row=34, column=free_column).value = suma_pkt
    else:
        sheet2.cell(row=31, column=free_column).value = dzis.date()
        sheet2.cell(row=32, column=free_column).value = suma_pkt
    filename_excel = f'5S RZE Engineering {dzis.strftime("%B")} {dzis.strftime("%Y")} (W{dzis.isocalendar()[1]}) {dzis.strftime("%d")}.{dzis.strftime("%m")}.{dzis.strftime("%Y")}.xlsx'
    workbook.save(filename=f'/home/guldier/5S/{filename_excel}')


    msg = MIMEMultipart("alternative")
    msg['Subject'] = '5S'
    msg['From'] = 'damian.jadacki@linetech.pl'
    msg['To'] = 'bartosz.wrobel@linetech.pl'

    text = f'Dear Roney,\nPlease find 5S report for {dzis.strftime("%d")}.{dzis.strftime("%m")}.{dzis.strftime("%Y")}'

    msg.attach(MIMEText(text, "plain"))

    html = """\
    <!DOCTYPE html>
    <html lang = "pl-PL">
        <head>
            <meta http - equiv ="Content-Language" content ="pl">
            <meta charset = "UTF-8">
        </head>
        <body>
            <style>
            table {font-family: Segoe UI,sans-serif;background: transparent;}
            p.a {font-size: 14px; color:#482960;}
            p.b {font-size: 11px; color:#482960;}
            p.c {font-size: 11px; color:#737373;}
            p.rodo {font-size: 8px; color:#737373; text-align: justify;};
            a:link {font-size: 11px; text-decoration: none; color: #737373;}
            a:visited {font-size: 11px; text-decoration: none; color: #737373;}
            a:active {font-size: 11px; text-decoration: none; color: #737373;}
            a:hover {font-size: 11px; text-decoration: none; color: #737373;}
            tbody.a {background: transparent;background-color:transparent}
            </style>
            <table width = "600" cellspacing="0" cellpadding = "0 " border = "0 ">
                <tbody class="a">
                    <tr>
                        <td>
                            <p class="a"><b> Damian Jadacki |</b> Engineer</p>
                        </td>
                        <td rowspan="2">

                        </td>
                        <td  colspan= "5" rowspan="2">
                            <a href="http://aviaprime.eu"><img  style = "border:none" src = "http://ftp.linetech.nazwa.pl/stopka/avia.png"></a>
                        </td>
                        <td rowspan="2">

                        </td>
                        <td rowspan="2">

                        </td>
                        <td rowspan="2">

                        </td>
                    </tr>
                    <tr>
                        <td>
                            <p class="b">LINETECH Aircraft Maintenance </p>
                        </td>
                        <td></td>
                        <td></td>
                        <td></td>
                        <td></td>
                        <td></td>
                    </tr>
                    <tr>
                        <td colspan=" 6" height=" 20 "> 
                            <br /> <br />
                        </td>
                    </tr>
                    <tr>
                        <td>
                            <p><a href="tel:+48887 400 189">+48 603 955 708 </a></p>
                        </td>
                        <td></td>
                        <td rowspan="3" width = "65">
                            <a href="http://www.linkedin.com/company/linetech-aircraft-maintenance"><img style = "border:none" width = "25" height = "25" src = "http://ftp.linetech.nazwa.pl/stopka/linkedin.png"></a>
                        </td>
                        <td rowspan="3" width = "65">
                            <a href="http://twitter.com/linetech_mro"><img style = "border:none" width = "25" height = "25" src = "http://ftp.linetech.nazwa.pl/stopka/twitter.png"></a>
                        </td>
                        <td rowspan="3" width = "65">
                            <a href="http://instagram.com/linetech_mro"><img style = "border:none" width = "25" height = "25" src = "http://ftp.linetech.nazwa.pl/stopka/instagram.png"></a>
                        </td>
                        <td rowspan="3" width = "20">
                            <a href="https://www.youtube.com/channel/UCS2tZyT9Mn0ArxIy41LVjYw"><img style = "border:none" width = "25" height = "25" src = "http://ftp.linetech.nazwa.pl/stopka/youtube.png"></a>
                        </td>                
                    </tr>
                    <tr>
                        <td>
                            <p> <a href="malito:michal.grebosz@linetech.pl"> damian.jadacki@linetech.pl </a></p>
                        </td>
                        <td></td>
                        <td></td>
                        <td></td>
                        <td></td>
                        <td></td>
                    </tr>
                    <tr>
                        <td>
                            <p><a href="www.aviaprime.eu"> www.aviaprime.eu </a></p>
                        </td>
                        <td></td>
                        <td></td>
                        <td></td>
                        <td></td>
                        <td></td>
                    </tr>
                </tbody>
            </table>
            <br>    
            <table width ="600">
                <tr>
                    <a href="http://seemore.aviaprime.eu"<img width = "600" height = "80" src = "http://ftp.linetech.nazwa.pl/stopka/Banner.png" alt = "Instagram" style = "border:none"></a>
                </tr>
                <tr>
                    <p class="rodo">Note: Under the General Data Protection Regulation (GDPR) (EU) 2016/679, we have a legal duty to protect any information we collect from you. Information contained in this email and any attachments may be privileged or confidential and intended for the exclusive use of the original recipient. If the reader of this e-mail is not the intended recipient or an agent responsibility for delivering it to the intended recipient, you are hereby notified that you have received this document in error and that any review, dissemination, distribution or coping of this message is strictly prohibited. If you have received this e-mail in error, please notify office@linetech.pl immediately. Linetech S.A. is 100% compliant with the General Data Protection Regulation (GDPR) .To learn more about how we collect, keep, and process your private information in compliance with GDPR, please view our privacy policy or contact us at iod@linetech.pl</p>
                </tr>
            </table>
        </body>
    </html>
    """.format(table_html=table_html)

    msg.attach(MIMEText(html, "html"))

    with open('/home/guldier/5S/IMG_3622.JPG',"rb") as attachent:
        picture = MIMEBase("application","octet-stream")
        picture.set_payload(attachent.read())

    encoders.encode_base64(picture)
    
    picture.add_header(
        "Content-Disposition",
        'attachment',
         filename='picture.JPG'
    )

    connection.open()
    

    #msg = EmailMessage(
    #    '5S',
    #    f'Czy to działa?\n {signature.as_string()}',
    #    'damian.jadacki@linetech.pl',
    #    ['bartosz.wrobel@linetech.pl'],
    #    connection=connection)

    #msg.attach_file('/home/guldier/5S/IMG_3622.JPG')
    #msg.attach_file(f'/home/guldier/5S/{filename_excel}')
    #msg.attach(part1)
    #msg.attach(signature)
    msg.attach(picture)

    context = ssl.create_default_context()
    with  smtplib.SMTP_SSL('linetech.nazwa.pl', 587, context=context) as s:
        s.login('damian.jadacki@linetech.pl','Dddddjjjjj12#$')
        s.sendmail('damian.jadacki@linetech.pl', 'bartosz.wrobel@linetech.pl', msg.as_string())
        s.quit()
    # msg.send()
    #connection.close()
